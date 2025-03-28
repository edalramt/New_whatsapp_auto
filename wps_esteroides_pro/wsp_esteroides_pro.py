import pandas as pd
import pyautogui as pg
import webbrowser as web
import subprocess
import time
import random
import pyperclip
import keyboard
import threading
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter  # Add this import

# Ruta de Chrome
ruta_chrome = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

# Ruta de los archivos PDF por facultad
pdf_paths = {
    "FCNM": r"FCNM_BROCHURE.pdf",
    "FIPA": r"FIPA_BROCHURE.pdf",
    "FIARN": r"FIARN_BROCHURE.pdf",
    "FIQ": r"FIQ_BROCHURE.pdf",
    "FIIS": r"FIIS_BROCHURE.pdf",
    "FIME": r"FIME_BROCHURE.pdf",
    "FIEE": r"FIEE_BROCHURE.pdf",
    "FCC": r"FCC_BROCHURE.pdf",
    "FCED": r"FCED_BROCURE.pdf",
    "FCE": {
        "Doctorado": r"FCE_DOCTORADO_BROCHURE.pdf",
        "Maestría": r"FCE_MAESTRIA_BROCHURE.pdf"
        },
    "FCA": r"FCA_BROCHURE.pdf",
    "FCS": {
        "Doctorado": r"FCS_DOCTORADO_BROCHURE.pdf",
        "Maestría": r"FCS_MAESTRIA_BROCHURE.pdf"
    }
}

# Función para normalizar números de teléfono
def normalizar_numero(numero):
    return numero.replace(" ", "")

# Función para verificar si un número es válido
def es_numero_valido(numero):
    return len(numero) == 9 and numero.isdigit()

# Función para esperar con variación aleatoria
def esperar_dinamico(tiempo_min, tiempo_max):
    tiempo_espera = random.uniform(tiempo_min, tiempo_max)
    time.sleep(tiempo_espera)

# Función para verificar si se ha presionado "Esc" o "0"
def verificar_tecla():
    global programa_activo
    while programa_activo:
        if keyboard.is_pressed('esc') or keyboard.is_pressed('0'):
            print("🛑 Programa detenido por el usuario.")
            programa_activo = False
            break
        time.sleep(0.1)

# Abrir el archivo Excel para actualizarlo
def abrir_archivo_excel(filename):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws.append(["celular", "nombre", "apellido", "programa", "name_programa", "facultad", "error"])
        wb.save(filename)
    return openpyxl.load_workbook(filename)
    

# Guardar una nueva fila en el archivo Excel
def guardar_nueva_fila(df, filename, nueva_fila):
    wb = abrir_archivo_excel(filename)
    ws = wb["Contactos"]
    for row in dataframe_to_rows(pd.DataFrame([nueva_fila], columns=df.columns), index=False, header=False):
        ws.append(row)
    wb.save(filename)

# Función para enviar mensaje
def enviar_mensaje(celular, nombre, apellido, programa, name_programa, facultad):
    saludos = [
        f"""👋 Hola {nombre} {apellido}!
Soy Edgar Ramos de la *Escuela de Posgrado de la UNAC*
🚀 *Tengo una noticia clave para tu desarrollo profesional!*""",
        f"""👋 Buenas tardes {nombre} {apellido}!
Soy Edgar Ramos de la *Escuela de Posgrado de la UNAC*
📢 Quiero que seas parte de esta gran oportunidad. *¡Atento a lo siguiente!*""",
        f"""👋 Hola {nombre} {apellido}!
Te habla Edgar Ramos desde la *Escuela de Posgrado de la UNAC*
💡 No dejes pasar esta información clave para tu futuro. *¡Mira esto!*"""
    ]

    mensaje_base = (
        f"""
📢 *ÚLTIMO DÌA PARA INSCRIBIRSE:*
{name_programa}
📅 Cierre de inscripciones: 25 de marzo

💰 *Costo de inscripción:* S/ 145
📂 *Incluye:* Carpeta de Postulante y Derecho de Inscripción

📅 *Fechas clave:*
    - Entrevista virtual: 26 y 27 de marzo
    - Resultados: 1-2 días después del examen
    - Inicio de clases: Primera semana de abril

"""
    )

    if facultad in ["FCS", "FIARN"]:
        mensaje_base += (
            "📍 *Modalidad de estudios:*\n"
            "    - 100% Virtual\n"
            "    - Fines de semana, de 8:00 a.m. a 8:00 p.m.\n"
            "\n"
        )
    else:
        mensaje_base += (
            "📍 *Modalidad de estudios:*\n"
            "    - 20% presencial y 80% virtual\n"
            "    - Presencial: 1 vez al mes (fines de semana, de 8:00 a.m. a 8:00 p.m.)\n"
            "\n"
        )

    if programa == "Doctorado":
        mensaje_base += (
            "⏳ *Duración del programa:* 6 semestres académicos\n"
            "💵 *Costo por semestre:* S/ ~~2500~~ S/ 2100\n"
            "\n"
        )
    elif programa == "Maestría":
        mensaje_base += (
            "⏳ *Duración del programa:* 3 semestres académicos\n"
            "💵 *Costo por semestre:* S/ ~~2500~~ S/ 2100\n"
            "\n"
        )

    mensaje_base += "🎯 ¡Inscríbete y prepárate para el siguiente nivel en tu formación profesional! 💼📖"

    pdf_mensaje = random.choice([
        "📎 Te adjunto el brochure con toda la información necesaria.",
        "📂 Aquí tienes el documento con la información detallada.",
        "🔍 Te envío el brochure oficial con los detalles del programa.",
        "📄 En el brochure adjunto encontrarás todos los detalles."
    ])
    consulta_mensaje = random.choice([
        """
🤝 Si tienes alguna duda o necesitas ayuda con tu inscripción, estoy aquí para apoyarte.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
✨ ¡Responde este mensaje y asegura tu inscripción hoy mismo!""",
        """
📢 *¡No pierdas esta oportunidad!*
Si tienes consultas, escríbeme y te ayudaré en lo que necesites.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
✅ Responde este mensaje y da el primer paso hacia tu futuro académico.""",
        """
📌 Estoy disponible para resolver cualquier duda y acompañarte en tu proceso de inscripción.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
🚀 ¡Escríbeme ahora y asegura tu cupo en la maestría!"""
    ])

    saludo = random.choice(saludos)

    inicio_tiempo = time.time()

    url = f"https://web.whatsapp.com/send?phone=+51{celular}"
    subprocess.Popen([ruta_chrome, url])
    esperar_dinamico(10, 16)
    
    pg.press("enter")

    pyperclip.copy(saludo)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")

    pyperclip.copy(mensaje_base)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")

    pyperclip.copy(pdf_mensaje)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 3)
    pg.press("enter")

    pg.click(745, 985)
    esperar_dinamico(2, 3)
    pg.click(779, 547)
    esperar_dinamico(1, 4)

    if facultad == "FCS" or facultad == "FCE":
        pdf_path = pdf_paths[facultad][programa]
    elif facultad == "FCED":
        pdf_path = pdf_paths["FCED"]["Doctorado"]
    else:
        pdf_path = pdf_paths[facultad]

    pyperclip.copy(pdf_path)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")
    esperar_dinamico(1, 6)
    pg.press("enter")
    esperar_dinamico(2, 4)

    pyperclip.copy(consulta_mensaje)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(1, 5)
    pg.press("enter")
    esperar_dinamico(2, 4)

    # Guardar el contacto en el DataFrame y escribir en el archivo Excel
    nuevo_contacto = [celular, nombre, apellido, programa, name_programa, facultad, ""]
    global contactos_enviados_df
    contactos_enviados_df = pd.concat([contactos_enviados_df, pd.DataFrame([nuevo_contacto], columns=contactos_enviados_df.columns)], ignore_index=True)
    guardar_nueva_fila(contactos_enviados_df, "contactos_enviados.xlsx", nuevo_contacto)

    fin_tiempo = time.time()
    duracion = fin_tiempo - inicio_tiempo
    print(f"✅ Mensaje enviado a {celular} en {duracion:.2f} segundos.")

    if duracion < 37:
        esperar_dinamico(37 - duracion, 42 - duracion)

    pg.hotkey("ctrl", "w")
    esperar_dinamico(4, 6)

# Función para guardar el DataFrame en una tabla de Excel
def guardar_contactos_enviados(df, filename):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Contactos')
            workbook = writer.book
            worksheet = writer.sheets['Contactos']
            
            # Definir la tabla con un estilo de color naranja
            tab = Table(displayName="ContactosEnviados", ref=f"A1:G{len(df)+1}")  # Ajustar rango
            style = TableStyleInfo(name="TableStyleMedium7", showFirstColumn=False,
                                    showLastColumn=True, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            
            # Ajustar automáticamente el ancho de las columnas
            for col_idx, col in enumerate(df.columns, start=1):  
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2  # Ajuste adicional
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length
            
    except PermissionError:
        print(f"❌ No se pudo escribir en '{filename}'. Verifique que el archivo no esté abierto y tenga los permisos necesarios.")

# Cargar contactos desde Excel
if not os.path.exists("preinscritos.xlsx"):
    print("Error: El archivo 'preinscritos.xlsx' no existe.")
    contactos_enviados_df = pd.DataFrame(columns=["celular", "nombre", "apellido", "programa", "name_programa", "facultad", "error"])
    guardar_contactos_enviados(contactos_enviados_df, "contactos_enviados.xlsx")
    exit()

try:
    data = pd.read_excel("preinscritos.xlsx")
except Exception as e:
    print(f"Error al cargar el archivo de Excel: {e}")
    contactos_enviados_df = pd.DataFrame(columns=["celular", "nombre", "apellido", "programa", "name_programa", "facultad", "error"])
    guardar_contactos_enviados(contactos_enviados_df, "contactos_enviados.xlsx")
    exit()

# Cargar contactos enviados previamente desde Excel
try:
    contactos_enviados_df = pd.read_excel("contactos_enviados.xlsx")
    if "celular" not in contactos_enviados_df.columns:
        raise KeyError("La columna 'celular' no existe en 'contactos_enviados.xlsx'")
    contactos_enviados = set(contactos_enviados_df["celular"].astype(str).apply(normalizar_numero))
except (FileNotFoundError, KeyError):
    contactos_enviados_df = pd.DataFrame(columns=["celular", "nombre", "apellido", "programa", "name_programa", "facultad", "error"])
    contactos_enviados = set()

# Variable para controlar la ejecución del programa
programa_activo = True

# Iniciar el hilo para verificar las teclas
hilo_tecla = threading.Thread(target=verificar_tecla)
hilo_tecla.start()

# Abrir WhatsApp Web en una nueva ventana
subprocess.Popen([ruta_chrome, "--new-window", "https://web.whatsapp.com"])
esperar_dinamico(11, 15)  # Esperar que cargue WhatsApp Web

for i in range(len(data)):
    if not programa_activo:
        break
    
    celular = normalizar_numero(str(data.loc[i, "cell"]))
    nombre = data.loc[i, "Nombre"]
    apellido = data.loc[i, "Apellido"]
    programa = data.loc[i, "Programa"]
    name_programa = data.loc[i, "Name_Programa"]
    facultad = data.loc[i, "Facultad"]
    
    if not es_numero_valido(celular):
        print(f"❌ Número inválido: {celular}. Saltando...")
        nuevo_contacto = [celular, nombre, apellido, programa, name_programa, facultad, "Número inválido"]
        contactos_enviados_df = pd.concat([contactos_enviados_df, pd.DataFrame([nuevo_contacto], columns=contactos_enviados_df.columns)], ignore_index=True)
        guardar_nueva_fila(contactos_enviados_df, "contactos_enviados.xlsx", nuevo_contacto)
        continue
    
    if celular in contactos_enviados:
        print(f"❌ {celular} ya fue contactado. Saltando...")
        nuevo_contacto = [celular, nombre, apellido, programa, name_programa, facultad, "Número repetido"]
        contactos_enviados_df = pd.concat([contactos_enviados_df, pd.DataFrame([nuevo_contacto], columns=contactos_enviados_df.columns)], ignore_index=True)
        guardar_nueva_fila(contactos_enviados_df, "contactos_enviados.xlsx", nuevo_contacto)
        continue

    try:
        enviar_mensaje(celular, nombre, apellido, programa, name_programa, facultad)
    except Exception as e:
        print(f"Error al enviar mensaje a {celular}: {e}")
        nuevo_contacto = [celular, nombre, apellido, programa, name_programa, facultad, str(e)]
        contactos_enviados_df = pd.concat([contactos_enviados_df, pd.DataFrame([nuevo_contacto], columns=contactos_enviados_df.columns)], ignore_index=True)
        guardar_nueva_fila(contactos_enviados_df, "contactos_enviados.xlsx", nuevo_contacto)

# Guardar el DataFrame final en el archivo Excel
guardar_contactos_enviados(contactos_enviados_df, "contactos_enviados.xlsx")





