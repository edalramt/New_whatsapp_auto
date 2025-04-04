import pandas as pd
import pyautogui as pg
import webbrowser as web
import subprocess
import time
import random
import pyperclip
import keyboard
import os
import threading
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Definir rutas de los archivos
archivo_preinscritos = os.path.join("auto_wsp","auto_wsp", "logistics", "Preinscritos.xlsx")
archivo_enviados = os.path.join("auto_wsp","auto_wsp", "logistics", "Contactos_enviados.xlsx")

# Verificar si el archivo de "Contactos_enviados.xlsx" existe
if os.path.exists(archivo_enviados):
    df_enviados = pd.read_excel(archivo_enviados)
else:
    df_enviados = pd.DataFrame(columns=["Nombre", "Apellido_Paterno", "Apellido_Materno", "DNI", "Facultad", "Programa", "Name_Programa", "Celular", "Correo", "Revision"])

# Manejar el caso en que "Preinscritos.xlsx" no exista
if not os.path.exists(archivo_preinscritos):
    print(f"Error: El archivo '{archivo_preinscritos}' no existe. Por favor, verifica la ruta o crea el archivo.")
    exit()

# Ruta de Chrome
ruta_chrome = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

# Ruta de los archivos PDF por facultad
pdf_paths = {
    "FCNM": r"FCNM_BROCHURE.pdf",
    "FIPA": r"FIPA_BROCHURE.pdf",
    "FIARN": r"FIARN_BROCHURE.pdf",
    "FIQ": r"FIQ_BROCHURE.pdf",
    "FIIS": r"FIIS_BROCHURE.pdf",
    "FIME": r"FIME_BROCHURE.pdf",
    "FIEE": r"FIEE_BROCHURE.pdf",
    "FCC": r"FCC_BROCHURE.pdf",
    "FCED": r"FCED_BROCURE.pdf",
    "FCE": {
        "Doctorado": r"FCE_DOCTORADO_BROCHURE.pdf",
        "Maestría": r"FCE_MAESTRIA_BROCHURE.pdf"
        },
    "FCA": r"FCA_BROCHURE.pdf",
    "FCS": {
        "Doctorado": r"FCS_DOCTORADO_BROCHURE.pdf",
        "Maestría": r"FCS_MAESTRIA_BROCHURE.pdf"
    }
}

# Función para verificar si se ha presionado "Esc" o "0"
def verificar_tecla():
    global programa_activo, programa_pausado
    programa_pausado = False
    while programa_activo:
        if keyboard.is_pressed('esc') or keyboard.is_pressed('0'):
            if not programa_pausado:
                print("⏸ Programa pausado. Presiona 'Esc' o '0' nuevamente para continuar.")
                programa_pausado = True
            else:
                print("▶ Programa reanudado.")
                programa_pausado = False
        time.sleep(0.1)
        
# Función para esperar con variación aleatoria
def esperar_dinamico(tiempo_min, tiempo_max):
    tiempo_espera = random.uniform(tiempo_min, tiempo_max)
    time.sleep(tiempo_espera)

# Función para normalizar números de teléfono
def normalizar_numero(numero):
    return numero.replace(" ", "")

# Función para verificar si un número es válido
def validar_numero(numero, enviados):
    numero = str(numero).replace(" ", "")
    if numero in enviados:
        return "Repetido"
    if numero.isdigit() and len(numero) == 9 and numero.startswith("9"):
        return "Válido"
    return "Inválido"

# Función para cargar datos desde Excel
def cargar_datos(archivo):
    if os.path.exists(archivo):
        return pd.read_excel(archivo)
    else:
        return pd.DataFrame(columns=["Nombre", "Apellido_Paterno", "Apellido_Materno", "DNI", "Facultad", "Programa", "Name_Programa", "Celular", "Correo", "Revision"])
    
def guardar_datos(df, archivo):
    try:
        df.to_excel(archivo, index=False, engine='openpyxl')

        # Aplicar estilo de tabla en Excel
        wb = load_workbook(archivo)
        ws = wb.active
        tabla = Table(displayName="ContactosEnviados", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
        
        # Estilo de tabla
        estilo = TableStyleInfo(name="TableStyleMedium7", showFirstColumn=False, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        # Ajustar automáticamente el ancho de las columnas
        for col_idx, col in enumerate(df.columns, start=1):  
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2  # Ajuste adicional
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length

        wb.save(archivo)
    except PermissionError:
        print(f"Error: No se puede guardar el archivo {archivo}. Asegúrate de que no esté abierto en otro programa.")
    except Exception as e:
        print(f"Error inesperado al guardar {archivo}: {e}")

def actualizar_datos(df_nuevos, archivo):
    """
    Actualiza un archivo Excel existente con nuevos datos.
    Si el archivo no existe, lo crea con los datos proporcionados.
    """
    try:
        # Si el archivo existe, cargarlo
        if os.path.exists(archivo):
            wb = load_workbook(archivo)
            ws = wb.active

            # Convertir los nuevos datos a filas
            for row in dataframe_to_rows(df_nuevos, index=False, header=False):
                ws.append(row)

        else:
            # Si el archivo no existe, crearlo con los datos nuevos
            with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
                df_nuevos.to_excel(writer, index=False, sheet_name='Contactos')

        # Guardar el archivo actualizado
        wb.save(archivo)
        print(f"✅ Archivo actualizado correctamente: {archivo}")

    except PermissionError:
        print(f"❌ Error: No se puede actualizar el archivo {archivo}. Asegúrate de que no esté abierto en otro programa.")
    except Exception as e:
        print(f"❌ Error inesperado al actualizar {archivo}: {e}")

def main():
    global programa_pausado
    df_preinscritos = cargar_datos(archivo_preinscritos)
    df_enviados = cargar_datos(archivo_enviados)
    enviados = df_enviados["Celular"].astype(str).values

    for _, row in df_preinscritos.iterrows():
        while programa_pausado:
            time.sleep(0.5)  # Esperar mientras el programa está pausado
        nombre = row["Nombre"]  
        apellido_paterno = row["Apellido_Paterno"]
        apellido_materno = row["Apellido_Materno"]
        dni = row["DNI"]
        facultad = row["Facultad"]
        programa = row["Programa"]
        name_programa = row["Name_Programa"]
        celular = str(row["Celular"])  # Convertir a string para evitar errores
        correo = row["Correo"]

        # Validar número
        estado = validar_numero(celular, enviados)

        # Si ya se envió antes, lo marcamos como "Repetido"
        if estado == "Válido":
            estado = enviar_mensaje(nombre, apellido_paterno, apellido_materno, dni, facultad, programa, name_programa, celular)

        # Evitar duplicados
        if not df_enviados["Celular"].isin([celular]).any():
            # Agregar los datos al DataFrame de enviados
            df_enviados = pd.concat([df_enviados, pd.DataFrame([{
                "Nombre": nombre,
                "Apellido_Paterno": apellido_paterno,
                "Apellido_Materno": apellido_materno,
                "DNI": dni,
                "Facultad": facultad,
                "Programa": programa,
                "Name_Programa": name_programa,
                "Celular": celular,
                "Correo": correo,
                "Revision": estado
            }])], ignore_index=True)

    # Guardar datos con estilo de tabla
    actualizar_datos(df_enviados, archivo_enviados)
    print("Proceso finalizado. Los mensajes fueron enviados y registrados.")

    
# Función para enviar mensajes por WhatsApp
def enviar_mensaje(Nombre, Apellido_Paterno, Apellido_Materno, Facultad, Programa, Name_Programa, Celular):
    saludos = [
        f"""👋 Hola {Nombre} {Apellido_Paterno} {Apellido_Materno}!
Soy Edgar Ramos de la *Escuela de Posgrado de la UNAC*
🚀 *Tengo una noticia clave para tu desarrollo profesional!*""",
        f"""👋 Buenas tardes {Nombre} {Apellido_Paterno} {Apellido_Materno}!
Soy Edgar Ramos de la *Escuela de Posgrado de la UNAC*
📢 Quiero que seas parte de esta gran oportunidad. *¡Atento a lo siguiente!*""",
        f"""👋 Hola {Nombre} {Apellido_Paterno} {Apellido_Materno}!
Te habla Edgar Ramos desde la *Escuela de Posgrado de la UNAC*
💡 No dejes pasar esta información clave para tu futuro. *¡Mira esto!*"""
    ]

    mensaje_base = (
        f"""
📢 *ÚLTIMO DÌA PARA INSCRIBIRSE:*
{Name_Programa}
📅 Cierre de inscripciones: 25 de marzo

💰 *Costo de inscripción:* S/ 145
📂 *Incluye:* Carpeta de Postulante y Derecho de Inscripción

📅 *Fechas clave:*
    - Entrevista virtual: 26 y 27 de marzo
    - Resultados: 1-2 días después del examen
    - Inicio de clases: Primera semana de abril

"""
    )

    if Facultad in ["FCS", "FIARN"]:
        mensaje_base += (
            "📍 *Modalidad de estudios:*\n"
            "    - 100% Virtual\n"
            "    - Fines de semana, de 8:00 a.m. a 8:00 p.m.\n"
            "\n"
        )
    else:
        mensaje_base += (
            "📍 *Modalidad de estudios:*\n"
            "    - 20% presencial y 80% virtual\n"
            "    - Presencial: 1 vez al mes (fines de semana, de 8:00 a.m. a 8:00 p.m.)\n"
            "\n"
        )

    if Programa == "Doctorado":
        mensaje_base += (
            "⏳ *Duración del programa:* 6 semestres académicos\n"
            "💵 *Costo por semestre:* S/ ~~2500~~ S/ 2100\n"
            "\n"
        )
    elif Programa == "Maestría":
        mensaje_base += (
            "⏳ *Duración del programa:* 3 semestres académicos\n"
            "💵 *Costo por semestre:* S/ ~~2500~~ S/ 2100\n"
            "\n"
        )

    mensaje_base += "🎯 ¡Inscríbete y prepárate para el siguiente nivel en tu formación profesional! 💼📖"

    pdf_mensaje = random.choice([
        "📎 Te adjunto el brochure con toda la información necesaria.",
        "📂 Aquí tienes el documento con la información detallada.",
        "🔍 Te envío el brochure oficial con los detalles del programa.",
        "📄 En el brochure adjunto encontrarás todos los detalles."
    ])
    consulta_mensaje = random.choice([
        """
🤝 Si tienes alguna duda o necesitas ayuda con tu inscripción, estoy aquí para apoyarte.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
✨ ¡Responde este mensaje y asegura tu inscripción hoy mismo!""",
        """
📢 *¡No pierdas esta oportunidad!*
Si tienes consultas, escríbeme y te ayudaré en lo que necesites.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
✅ Responde este mensaje y da el primer paso hacia tu futuro académico.""",
        """
📌 Estoy disponible para resolver cualquier duda y acompañarte en tu proceso de inscripción.
📩 *Correo:* posgrado.admision@unac.edu.pe
📞 *WhatsApp:* 900969591\n
🚀 ¡Escríbeme ahora y asegura tu cupo en la maestría!"""
    ])

    saludo = random.choice(saludos)

    inicio_tiempo = time.time()

    url = f"https://web.whatsapp.com/send?phone=+51{Celular}"
    subprocess.Popen([ruta_chrome, url])
    esperar_dinamico(10, 16)
    
    pg.press("enter")

    pyperclip.copy(saludo)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")

    pyperclip.copy(mensaje_base)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")

    pyperclip.copy(pdf_mensaje)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 3)
    pg.press("enter")

    pg.click(745, 985)
    esperar_dinamico(2, 3)
    pg.click(779, 547)
    esperar_dinamico(1, 4)

    if Facultad == "FCS" or Facultad == "FCE":
        pdf_path = pdf_paths[Facultad][Programa]
    elif Facultad == "FCED":
        pdf_path = pdf_paths["FCED"]["Doctorado"]
    else:
        pdf_path = pdf_paths[Facultad]
    pyperclip.copy(pdf_path)
    
    pg.hotkey("ctrl", "v")
    esperar_dinamico(2, 4)
    pg.press("enter")
    esperar_dinamico(1, 6)
    pg.press("enter")
    esperar_dinamico(2, 4)

    pyperclip.copy(consulta_mensaje)
    pg.hotkey("ctrl", "v")
    esperar_dinamico(1, 5)
    pg.press("enter")
    esperar_dinamico(2, 4)

    fin_tiempo = time.time()
    duracion = fin_tiempo - inicio_tiempo
    print(f"✅ Mensaje enviado a {Celular} en {duracion:.2f} segundos.")

    if duracion < 37:
        esperar_dinamico(37 - duracion, 42 - duracion)

    pg.hotkey("ctrl", "w")
    esperar_dinamico(4, 6)
    return "Enviado"

# Ejecutar el script
if __name__ == "__main__":
    main()